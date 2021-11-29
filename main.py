from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook
import time


account = input('Please enter the Hike account ID: ')
username = input('Please enter the Hike username: ')
password = input('Please enter the password: ')
filePath = input("Where's the file? ")
logInButton = "/html[1]/body[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[2]/form[1]/div[4]/button[1]"
salesSectionOnDashboard = "/html[1]/body[1]/div[1]/div[3]/div[2]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[2]/div[1]/div[1]"
productsInMenu = "/html[1]/body[1]/div[1]/div[3]/div[1]/div[1]/div[1]/ul[1]/li[11]/a[1]"
productsInSubMenu = "/html[1]/body[1]/div[1]/div[3]/div[1]/div[1]/div[1]/ul[1]/li[11]/ul[1]/li[1]/a[1]"
filterOnProductsPage = "/html[1]/body[1]/div[1]/div[3]/div[2]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[2]/div[1]/div[1]/div[1]/div[1]/div[1]/input[1]"
firstProductMatching = "/html[1]/body[1]/div[1]/div[3]/div[2]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[2]/div[1]/div[1]/div[1]/ul[1]/li[1]/ul[2]/li[1]/a[1]"
filteredProduct = "/html[1]/body[1]/div[1]/div[3]/div[2]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[3]/div[1]/div[1]/div[1]/table[1]/tbody[1]/tr[1]/td[2]/span[1]"
productNameField = "/html[1]/body[1]/div[1]/div[3]/div[2]/div[1]/div[1]/div[1]/form[1]/div[1]/div[4]/div[2]/div[2]/div[2]/div[1]/input[1]"
sameRetailPrice = "/html[1]/body[1]/div[1]/div[3]/div[2]/div[1]/div[1]/div[1]/form[1]/div[1]/div[4]/div[2]/div[3]/div[2]/div[1]/div[2]"
saveButtonTop = "/html[1]/body[1]/div[1]/div[3]/div[2]/div[1]/div[1]/div[1]/form[1]/div[1]/div[2]/button[2]/span[1]"
xButtonForFilter = "/html[1]/body[1]/div[1]/div[3]/div[2]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[2]/div[1]/div[1]/div[2]/div[1]/span[1]/a[1]/i[1]"

driver = webdriver.Chrome(ChromeDriverManager().install())
# Log into the store and navigate to the product listing page
driver.get("https://my.hikeup.com/")
driver.set_window_size(1920, 1080)
driver.find_element(By.NAME, "tenancyName").send_keys(account)
driver.find_element(By.NAME, "usernameOrEmailAddress").send_keys(username)
driver.find_element(By.NAME, "password").send_keys(password)
driver.find_element(By.XPATH, logInButton).click()
wait = WebDriverWait(driver, 60)
element = wait.until(EC.visibility_of_element_located((By.XPATH, salesSectionOnDashboard)))
driver.find_element(By.XPATH, productsInMenu).click()
driver.find_element(By.XPATH, productsInSubMenu).click()
wait.until(EC.visibility_of_element_located((By.XPATH, filterOnProductsPage)))

def resave_product(sku):
    #filter for the product by SKU
    driver.find_element(By.XPATH, filterOnProductsPage).click()
    driver.find_element(By.XPATH, filterOnProductsPage).send_keys(sku)
    wait.until(EC.visibility_of_element_located((By.XPATH, firstProductMatching)))
    driver.find_element(By.XPATH, firstProductMatching).click()
    wait.until(EC.visibility_of_element_located((By.XPATH, filteredProduct)))
    #click into the product profile
    driver.find_element(By.XPATH,filteredProduct).click()
    wait.until(EC.visibility_of_element_located((By.XPATH, productNameField)))
    time.sleep(3)
    actions = ActionChains(driver)
    actions.move_to_element(driver.find_element(By.XPATH, sameRetailPrice)).perform()
    driver.find_element(By.XPATH, saveButtonTop).click()
    #clear the filter
    wait.until(EC.visibility_of_element_located((By.XPATH, xButtonForFilter)))
    driver.find_element(By.XPATH, xButtonForFilter).click()
    time.sleep(5)

#read SKU List from the file
wb = load_workbook(filePath)
ws = wb.active
column = ws['A']
skuList = [column[x].value for x in range(len(column))]
#resave the products
for sku in skuList:
    resave_product(sku)
    print(f"{sku} has been resaved")

driver.close()
